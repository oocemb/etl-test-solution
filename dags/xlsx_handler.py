from airflow.providers.postgres.hooks.postgres import PostgresHook
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from settings import logger, settings


def insert_overdue_raw_data():
    pg_hook = PostgresHook(postgres_conn_id="pg_localhost")
    pg_conn = pg_hook.get_conn()
    pg_curs = pg_conn.cursor()

    # get from xcom list path to handle
    path_to_file = settings.PATH_TO_FILES_DOCKER + "Просрочено (06.09.2022).xlsx"
    logger.info(f"Handle {path_to_file}")
    book: Workbook = load_workbook(filename=path_to_file, data_only=True)
    sheets: Worksheet = book.active
    insert_list = []

    for row_cells in sheets.iter_rows(min_row=6):
        insert_list.append([cell.value for cell in row_cells if cell.value is not None])

    args_str = ','.join(
        pg_curs.mogrify("(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", x).decode() for x in insert_list)

    pg_curs.execute("INSERT INTO overdue(subject, mo, inn, status, type_out, gtin, series, count_doses_in_pkg, "
                    "count_packages, count_doses_sum, expiration_date, overdue_day) VALUES " + args_str)
    pg_conn.commit()
    pg_conn.close()
    logger.info("Success load batch data")
